import re
import math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


LAB_INPUT_FILE = "input.csv"
SOIL_PSRG_FILE = "20250116_PSRG.xlsx"
OUTPUT_FILE = "grouped_lab_detections.xlsx"


def is_detection(result):
    if pd.isna(result):
        return False

    result_str = str(result).strip()

    if result_str == "":
        return False

    if result_str.startswith("<"):
        return False

    try:
        float(result_str)
        return True
    except ValueError:
        return False


def normalize_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def format_limit_value(value):
    if pd.isna(value) or value == "":
        return ""

    try:
        num = float(value)
        if math.isclose(num, round(num)):
            return str(int(round(num)))
        return f"{num:.10f}".rstrip("0").rstrip(".")
    except Exception:
        return str(value).strip()


def safe_numeric(value):
    if pd.isna(value):
        return None

    value_str = str(value).strip()

    if value_str == "":
        return None

    if value_str.startswith("<"):
        return None

    value_str = value_str.replace(",", "")

    try:
        return float(value_str)
    except ValueError:
        return None


def combine_result_and_qualifier(result, qualifier):
    result_str = normalize_text(result)
    qualifier_str = normalize_text(qualifier)

    if qualifier_str:
        return f"{result_str} {qualifier_str}"
    return result_str


def parse_sample_id_parts(sample_id):
    if pd.isna(sample_id):
        return "", "", ""

    sample_str = str(sample_id).strip()
    match = re.match(r"^(.*?)\s*\(([^()]*)\)\s*$", sample_str)

    if not match:
        return sample_str, "", ""

    base = match.group(1).strip()
    inside = match.group(2).strip()

    if re.match(r"^\d+(\.\d+)?\s*-\s*\d+(\.\d+)?$", inside):
        return base, inside, ""

    return f"{base}\n({inside})", "", inside


def categorize_sample(sample_id):
    if pd.isna(sample_id):
        return "Other"

    sample_str = str(sample_id).strip().upper()

    if sample_str.startswith("TMW") or sample_str.startswith("MW"):
        return "Groundwater"

    if sample_str.startswith("SB"):
        return "Soil"

    if sample_str.startswith("DUP"):
        match = re.match(r"^(.*?)\s*\(([^()]*)\)\s*$", sample_str)
        if match:
            parent = match.group(2).strip().upper()
            if parent.startswith("SB"):
                return "Soil"
            if parent.startswith("TMW") or parent.startswith("MW"):
                return "Groundwater"
        return "Other"

    return "Other"


def format_sample_date(value):
    if pd.isna(value):
        return ""

    try:
        dt = pd.to_datetime(value)
        return f"{dt.month}/{dt.day}/{dt.year}"
    except Exception:
        return str(value).strip()


def clean_dataframe(df):
    df = df.copy()
    df.columns = df.columns.str.strip()

    needed_columns = [
        "Sample Id",
        "CAS Number",
        "Analyte Description",
        "Result",
        "Qualifier",
        "Sampled DateTime",
        "Method",
        "Unit of Measure"
    ]

    existing = [c for c in needed_columns if c in df.columns]
    df = df[existing].copy()

    df = df[df["Analyte Description"].astype(str).str.strip().str.upper() != "MOISTURE"].copy()

    df["Sample Category"] = df["Sample Id"].apply(categorize_sample)
    df["Is Detection"] = df["Result"].apply(is_detection)
    df["Numeric Result"] = df["Result"].apply(safe_numeric)

    df["Display Result"] = df.apply(
        lambda row: combine_result_and_qualifier(row.get("Result"), row.get("Qualifier")),
        axis=1
    )

    parsed = df["Sample Id"].apply(parse_sample_id_parts)
    df["Display Sample Id"] = parsed.apply(lambda x: x[0])
    df["Sample Depth"] = parsed.apply(lambda x: x[1])
    df["Parent Sample"] = parsed.apply(lambda x: x[2])

    df["Display Sample Date"] = df["Sampled DateTime"].apply(format_sample_date)

    return df


def load_soil_psrg(psrg_file):
    psrg_df = pd.read_excel(psrg_file, header=10)
    psrg_df.columns = psrg_df.columns.str.strip()

    rename_map = {
        "Chemical Name": "Analyte Description",
        "Residential Health Based PSRG (mg/kg)": "Residential PSRG",
        "Industrial/Commercial Health Based PSRG (mg/kg)": "Industrial PSRG",
        "Protection of Groundwater PSRG (mg/kg)": "Groundwater Protection PSRG",
    }

    psrg_df = psrg_df.rename(columns=rename_map)

    needed = [
        "CAS #",
        "Analyte Description",
        "Residential PSRG",
        "Industrial PSRG",
        "Groundwater Protection PSRG",
    ]

    psrg_df = psrg_df[needed].copy()

    psrg_df["CAS #"] = psrg_df["CAS #"].astype(str).str.strip()
    psrg_df["Analyte Description"] = psrg_df["Analyte Description"].astype(str).str.strip()

    return psrg_df


def build_group_table(group_df, soil_psrg_df=None):
    detected_analytes = (
        group_df.groupby("Analyte Description")["Is Detection"]
        .any()
        .reset_index()
    )

    detected_analytes = detected_analytes.loc[
        detected_analytes["Is Detection"],
        "Analyte Description"
    ]

    filtered_df = group_df[group_df["Analyte Description"].isin(detected_analytes)].copy()

    if filtered_df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    sample_order = filtered_df["Display Sample Id"].drop_duplicates().tolist()

    pivot_df = filtered_df.pivot_table(
        index=["CAS Number", "Analyte Description"],
        columns="Display Sample Id",
        values="Display Result",
        aggfunc="first"
    ).reindex(columns=sample_order)

    detect_status_df = filtered_df.pivot_table(
        index=["CAS Number", "Analyte Description"],
        columns="Display Sample Id",
        values="Is Detection",
        aggfunc="first"
    ).reindex(columns=sample_order)

    numeric_df = filtered_df.pivot_table(
        index=["CAS Number", "Analyte Description"],
        columns="Display Sample Id",
        values="Numeric Result",
        aggfunc="first"
    ).reindex(columns=sample_order)

    depth_map = (
        filtered_df.groupby("Display Sample Id")["Sample Depth"]
        .first()
        .reindex(sample_order)
    )

    date_map = (
        filtered_df.groupby("Display Sample Id")["Display Sample Date"]
        .first()
        .reindex(sample_order)
    )

    result_table = pivot_df.reset_index()
    result_table = result_table.rename(columns={"Analyte Description": "Analyte", "CAS Number": "CAS Number"})

    numeric_summary = (
        filtered_df.groupby(["CAS Number", "Analyte Description"])["Numeric Result"]
        .max()
        .reset_index()
        .rename(columns={"Numeric Result": "Highest Concentration"})
    )

    result_table = result_table.merge(
        numeric_summary,
        left_on=["CAS Number", "Analyte"],
        right_on=["CAS Number", "Analyte Description"],
        how="left"
    ).drop(columns=["Analyte Description"])

    if soil_psrg_df is not None:
        result_table = result_table.merge(
            soil_psrg_df,
            left_on=["CAS Number", "Analyte"],
            right_on=["CAS #", "Analyte Description"],
            how="left"
        )

        result_table = result_table.drop(columns=["CAS #", "Analyte Description"])

        for col in [
            "Highest Concentration",
            "Residential PSRG",
            "Industrial PSRG",
            "Groundwater Protection PSRG"
        ]:
            if col in result_table.columns:
                result_table[col] = result_table[col].apply(format_limit_value)

    header_rows = []

    if depth_map.fillna("").astype(str).str.strip().any():
        depth_values = ["", "Sample Depth (ft bgs)"] + depth_map.fillna("").tolist()
        while len(depth_values) < len(result_table.columns):
            depth_values.append("")
        header_rows.append(depth_values)

    date_values = ["", "Sampling Date"] + date_map.fillna("").tolist()
    while len(date_values) < len(result_table.columns):
        date_values.append("")
    header_rows.append(date_values)

    final_columns = ["CAS Number", "Analyte"] + sample_order

    extra_cols = [col for col in result_table.columns if col not in final_columns]
    final_columns += extra_cols

    result_table = result_table[final_columns]

    if header_rows:
        header_rows_df = pd.DataFrame(header_rows, columns=final_columns)
        final_df = pd.concat([header_rows_df, result_table], ignore_index=True)
    else:
        final_df = result_table

    return final_df, detect_status_df, numeric_df


def apply_output_formatting(output_file, sheet_detection_maps, sheet_numeric_maps):
    wb = load_workbook(output_file)

    gray_font = Font(color="808080")
    shaded_fill = PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]

        header_map = {header: idx for idx, header in enumerate(headers, start=1) if header is not None}

        if sheet_name not in sheet_detection_maps:
            continue

        detect_status_df = sheet_detection_maps[sheet_name]
        numeric_df = sheet_numeric_maps[sheet_name]

        sample_columns = [col for col in detect_status_df.columns]

        first_analyte_row = 2
        for row_num in range(2, ws.max_row + 1):
            analyte_val = ws.cell(row=row_num, column=header_map["Analyte"]).value
            if analyte_val not in ("Sample Depth (ft bgs)", "Sampling Date"):
                first_analyte_row = row_num
                break

        for excel_row in range(first_analyte_row, ws.max_row + 1):
            cas_number = ws.cell(row=excel_row, column=header_map["CAS Number"]).value
            analyte = ws.cell(row=excel_row, column=header_map["Analyte"]).value

            key = (cas_number, analyte)
            if key not in detect_status_df.index:
                continue

            res_limit = None
            ind_limit = None
            gw_limit = None

            if "Residential PSRG" in header_map:
                res_val = ws.cell(row=excel_row, column=header_map["Residential PSRG"]).value
                res_limit = safe_numeric(res_val)

            if "Industrial PSRG" in header_map:
                ind_val = ws.cell(row=excel_row, column=header_map["Industrial PSRG"]).value
                ind_limit = safe_numeric(ind_val)

            if "Groundwater Protection PSRG" in header_map:
                gw_val = ws.cell(row=excel_row, column=header_map["Groundwater Protection PSRG"]).value
                gw_limit = safe_numeric(gw_val)

            for sample_id in sample_columns:
                if sample_id not in header_map:
                    continue

                col_idx = header_map[sample_id]
                cell = ws.cell(row=excel_row, column=col_idx)

                is_detect = detect_status_df.loc[key, sample_id]
                numeric_result = numeric_df.loc[key, sample_id]

                if pd.isna(is_detect):
                    continue

                if is_detect is False:
                    cell.font = Font(color="808080")
                    continue

                bold = False
                underline = None
                fill = None

                if numeric_result is not None:
                    if res_limit is not None and numeric_result > res_limit:
                        bold = True

                    if ind_limit is not None and numeric_result > ind_limit:
                        fill = shaded_fill

                    if gw_limit is not None and numeric_result > gw_limit:
                        underline = "single"

                cell.font = Font(bold=bold, underline=underline)
                if fill is not None:
                    cell.fill = fill

    wb.save(output_file)


def main():
    df = pd.read_csv(LAB_INPUT_FILE, encoding="latin1")
    df = clean_dataframe(df)

    soil_psrg_df = load_soil_psrg(SOIL_PSRG_FILE)

    sheet_detection_maps = {}
    sheet_numeric_maps = {}

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        for category in df["Sample Category"].dropna().unique():
            group_df = df[df["Sample Category"] == category].copy()

            if group_df.empty:
                continue

            psrg_to_use = soil_psrg_df if category == "Soil" else None

            output_table, detect_status_df, numeric_df = build_group_table(group_df, psrg_to_use)

            if output_table.empty:
                continue

            sheet_name = str(category)[:31]
            output_table.to_excel(writer, sheet_name=sheet_name, index=False)

            sheet_detection_maps[sheet_name] = detect_status_df
            sheet_numeric_maps[sheet_name] = numeric_df

    apply_output_formatting(OUTPUT_FILE, sheet_detection_maps, sheet_numeric_maps)

    print(f"Done. Output saved to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()